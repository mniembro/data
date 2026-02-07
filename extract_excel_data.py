import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from PIL import Image

def filter_columns(df, columns_to_keep):
    """
    Filters data frame to keep only specified columns.
    
    parameters
    - df: pandas DataFrame original
    - columns: List of column names to keep

    Returns:
    - A new DataFrame with only the specified columns
    """
    try:
        # Filtrar solo las columnas que existen en el DataFrame
        valid_columns = [col for col in columns_to_keep if col in df.columns]

        # Verificar si hay columnas no encontradas
        not_found_columns = [col for col in columns_to_keep if col not in df.columns]
        if not_found_columns:
            print(f"warning: following columns were ignored: {', '.join(not_found_columns)}")

        if not valid_columns:
            print("Error: Specified columns not found in the DataFrame.")
            return None

        return df[valid_columns]
    except Exception as e:
        print(f"Error filtering columns: {e}")
        return None


def read_excel(file_path, sheet=0, header_line=0, column_names=None, extract_all_sheets=True):
    """
    Function to read an Excel file and return a pandas DataFrame.

    Parameters:
    - file_path: Full path to the Excel file
    - sheet: Name or index of the sheet to read (default is the first sheet)
    - header_line: Row number (0-indexed) where column names are located
    - nombres_columnas: Optional list of custom column names
    - extraer_todas_hojas: If True, combines all sheets into a single DataFrame

    Returns:
    - A pandas DataFrame with the Excel data
    """
    try:
        if extract_all_sheets:
            xls = pd.ExcelFile(file_path)
            all_sheets = []
            
            for sheet_name in xls.sheet_names:
                df = pd.read_excel(
                    xls,
                    sheet_name=sheet_name,
                    header=header_line
                )
                
                df.columns = df.columns.str.strip()

                df['Original_sheet'] = sheet_name
                
                all_sheets.append(df)
            
            # Combinar todas las hojas
            if all_sheets:
                combined_df = pd.concat(all_sheets, ignore_index=True)
                print("\nAll sheets read successfully and combined.")
                print(f"Total sheets processed: {len(all_sheets)}")
                print(f"Total rows combined: {len(combined_df)}")
                print("\nAvailable columns:")
                print(combined_df.columns.tolist())
                return combined_df
            return None
            
        else:
            # Leer solo la hoja especificada
            df = pd.read_excel(
                file_path,
                sheet_name=sheet,
                header=header_line
            )
            
            # Limpiar nombres de columnas
            df.columns = df.columns.str.strip()
            
            print("\nData read successfully from a single sheet.")
            print(f"Total rows: {len(df)}")
            print("\nAvailable columns:")
            print(df.columns.tolist())
            return df
            
    except Exception as e:
        print(f"Error reading the file: {e}")
        return None

def extract_excel_sheet_images(file_path, sheet=0, output_folder='extracted_images'):
    """
    Extract embedded images from an Excel file.

    Parameters:
    - file_path: Path to the Excel file
    - sheet: Name or index of the sheet (default is the first sheet)
    - output_folder: Folder where the extracted images will be saved

    Returns:
    - List of paths to the saved images
    """
    try:
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
            print(f"Carpeta '{output_folder}' creada correctamente.")
        
        wb = load_workbook(file_path)
        if isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]


        image_loader = SheetImageLoader(ws)
        saved_images = []
        for row in ws.iter_rows():
            for cell in row:
                if image_loader.image_in(cell.coordinate):
                    try:
                        # Get image
                        image = image_loader.get(cell.coordinate)
                        
                        # Get left cell text for naming
                        col_letra = cell.column_letter
                        fila = cell.row
                        if cell.column > 1: 
                            col_izq = chr(ord(col_letra) - 1)
                            celda_izq = ws[f'{col_izq}{fila}']
                            texto_celda = str(celda_izq.value).strip() if celda_izq.value else ""
                        else:
                            texto_celda = ""
                        if texto_celda:
                            # strip
                            safe_name = "".join(c if c.isalnum() or c in (' ', '-', '_') else '_' for c in texto_celda)
                            safe_name = safe_name[:50]  # Limitar la longitud del nombre
                            file_name = f"{safe_name}.png"
                        else:
                            file_name = f"imagen_{cell.coordinate}.png"
                            
                        image_path = os.path.join(output_folder, file_name)
                        
                        counter = 1
                        base_name, ext = os.path.splitext(image_path)
                        while os.path.exists(image_path):
                            image_path = f"{base_name}_{counter}{ext}"
                            counter += 1

                        image.save(image_path)
                        saved_images.append(image_path)
                        print(f"Image saved as: {os.path.basename(image_path)}")

                    except Exception as e:
                        print(f"Error processing image: {cell.coordinate}: {e}")

        if not saved_images:
            print("No images found in the Excel file.")

        return saved_images

    except Exception as e:
        print(f"Error extracting images: {e}")
        return []


def save_excel(df, output_file_name):
    """
    Function to save a DataFrame to an Excel file

    Parameters:
    - df: Pandas DataFrame to save
    - output_file_name: Name of the output file
    """
    try:
        df.to_excel(output_file_name, index=False)
        print(f"Data successfully saved to {output_file_name}")
    except Exception as e:
        print(f"Error saving file: {e}")

import argparse

def list_sheets(file_path):
    """Displays all available sheets in the Excel file."""
    try:
        # Use openpyxl to get sheet names
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        print("\nAvailable sheets in the file:")
        for i, sheet in enumerate(sheets):
            print(f"{i}: {sheet}")
        return sheets
    except Exception as e:
        print(f"Error reading sheets from file: {e}")
        return []

def parse_arguments():
    """Configures and parses command line arguments."""
    parser = argparse.ArgumentParser(description='Extract images from an Excel file.')
    parser.add_argument('--archivo', type=str, required=True,
                       help='Path to the Excel file')
    parser.add_argument('--hoja', type=int, default=0,
                       help='Sheet number (0 for the first sheet)')
    parser.add_argument('--extraer_todas_hojas', action='store_true',
                       help='Combine all sheets into a single DataFrame')
    parser.add_argument('--carpeta', type=str, default='imagenes_extraidas',
                       help='Name of the output folder')
    parser.add_argument('--extract_images', action='store_true',
                       help='Extract images from the Excel file')
    parser.add_argument('--listar_hojas', action='store_true',
                       help='Show the list of available sheets')

    return parser.parse_args()

# Example usage
if __name__ == "__main__":
    args = parse_arguments()

    # Show list of sheets if requested
    if args.listar_hojas:
        list_sheets(args.archivo)
        exit()

    # Extract images with the provided arguments
    if args.extract_images:
        sheets = list_sheets(args.archivo)
        for sheet in range(len(sheets)):
            images = extract_excel_sheet_images(
                file_path=args.archivo,
                sheet=sheet,
                output_folder=args.carpeta
            )
            print(f"\nTotal extracted images: {len(images)}")

    data = read_excel(args.archivo, sheet=args.hoja, extract_all_sheets=args.extraer_todas_hojas)

    if data is not None:
        # Show available columns
        print("\nAvailable columns in the data:")
        print(data.columns.tolist())

        # Specify the columns to keep
        columns_to_keep = ['Codigo de Barras', 'Articulo', "Empaque", "Origen_Hoja"]

        if columns_to_keep:
            # Filter the columns
            filtered_data = filter_columns(data, columns_to_keep)

            if filtered_data is not None:
                print("\nData after filtering columns (first 5 rows):")
                print(filtered_data.head())
                print(f"\nTotal rows: {len(filtered_data)}")

                # Save the filtered data as CSV
                output_name = 'datos_combinados.csv'
                filtered_data.to_csv(output_name, index=False, encoding='utf-8-sig')
                print(f"\nData saved to: {output_name}")